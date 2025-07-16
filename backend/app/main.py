from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import uuid
import os
import json
from dotenv import load_dotenv

from .chatbot import (
    build_json_system_prompt,
    build_insights_prompt,
    build_qna_prompt,
    get_structured_llm_response
)

load_dotenv()

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_methods=["*"],
    allow_headers=["*"],
)

class ChatRequest(BaseModel):
    step: int
    user_state: dict

@app.post("/chat")
def chat(req: ChatRequest):
    print("\n📨 [DEBUG] Received POST /chat payload:")
    print(json.dumps(req.dict(), indent=2))

    step = req.step
    state = req.user_state
    name = state.get("name", "")
    goal = state.get("goal", "")
    income = state.get("income", {})
    expenses = state.get("expenses", {})
    qna = state.get("qna", "")

    STEPS = {
        "QNA": 16,
        "DOWNLOAD": 17
    }

    if step == STEPS["QNA"]:
        messages = [
            {"role": "system", "content": build_json_system_prompt()},
            {"role": "user", "content": build_qna_prompt(name, goal, income, expenses, qna)},
        ]
        print("\n🧠 [DEBUG] Messages to OpenAI for QNA:")
        for m in messages:
            print(f"  {m['role']}: {m['content'][:200]}" + ("..." if len(m['content']) > 200 else ""))
        try:
            llm_response = get_structured_llm_response(messages)
            print("\n📥 [DEBUG] Raw LLM JSON response:", llm_response)
            result = json.loads(llm_response)
            if not all(k in result for k in ["summary", "tips"]):
                raise ValueError("Response missing required fields.")
            return {"response": result}
        except Exception as e:
            print("❌ [DEBUG] LLM JSON parse error (QNA):", str(e))
            return {"response": {
                "summary": "Sorry, there was a problem generating advice.",
                "tips": []
            }}

    return {"response": "Step does not require backend logic (handled in frontend)."}

def style_cell(cell, bold=False):
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

def write_section(ws, title, rows):
    ws.append([])
    ws.append([title])
    style_cell(ws.cell(row=ws.max_row, column=1), bold=True)
    for row in rows:
        ws.append(row)
        for col in range(1, len(row) + 1):
            style_cell(ws.cell(row=ws.max_row, column=col))

@app.post("/export-budget")
def export_budget(data: dict):
    print("\n🟢 [DEBUG] Exporting budget with data:")
    print(json.dumps(data, indent=2))
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Summary"

    write_section(ws, "Financial Goal", [[data.get("goal", "N/A")]])

    income_rows = [["Source", "Amount ($)"]] + [[src, amt] for src, amt in data.get("income", {}).items()]
    write_section(ws, "Income Sources", income_rows)

    expense_rows = [["Category", "Amount ($)"]] + [[cat, amt] for cat, amt in data.get("expenses", {}).items()]
    write_section(ws, "Expense Categories", expense_rows)

    total_income = sum(data.get("income", {}).values())
    total_expenses = sum(data.get("expenses", {}).values())
    write_section(ws, "Summary", [
        ["Total Income", total_income],
        ["Total Expenses", total_expenses],
        ["Net Balance", total_income - total_expenses]
    ])

    name = data.get("name", "")
    goal = data.get("goal", "")
    income = data.get("income", {})
    expenses = data.get("expenses", {})

    messages = [
        {"role": "system", "content": build_json_system_prompt()},
        {"role": "user", "content": build_insights_prompt(name, goal, income, expenses)},
    ]
    print("\n🧠 [DEBUG] Messages to OpenAI for spreadsheet insights:")
    for m in messages:
        print(f"  {m['role']}: {m['content'][:200]}" + ("..." if len(m['content']) > 200 else ""))
    try:
        llm_json = get_structured_llm_response(messages)
        print("\n📥 [DEBUG] Raw LLM JSON response (export):", llm_json)
        result = json.loads(llm_json)
        summary = result.get("summary", "")
        tips = result.get("tips", [])
        write_section(ws, "AI Budget Summary", [[summary]])
        if tips:
            write_section(ws, "AI Budget Tips", [[f"• {tip}"] for tip in tips])
    except Exception as e:
        print("❌ [DEBUG] LLM JSON parse error (export):", str(e))
        write_section(ws, "AI Budget Summary", [["Sorry, couldn't generate insights for your spreadsheet."]])

    os.makedirs("downloads", exist_ok=True)
    filename = f"budget_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join("downloads", filename)
    wb.save(filepath)

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
